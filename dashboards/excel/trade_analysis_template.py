#!/usr/bin/env python3
"""
Excel Template Generator for Trade Analysis
Creates comprehensive Excel workbooks with financial modeling, pivot tables, and scenario analysis.
"""

import pandas as pd
import sqlite3
import numpy as np
from pathlib import Path
import sys
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

# Add the project root to the Python path
project_root = Path(__file__).parent.parent.parent
sys.path.append(str(project_root))

class ExcelTemplateGenerator:
    """Generate Excel templates for trade analysis."""
    
    def __init__(self):
        self.db_path = project_root / "data" / "sql" / "trade_analysis.db"
        self.output_dir = project_root / "dashboards" / "excel"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
    def generate_template(self):
        """Generate comprehensive Excel template."""
        print("Generating Excel template...")
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self.create_dashboard_sheet(wb)
        self.create_trade_data_sheet(wb)
        self.create_economic_indicators_sheet(wb)
        self.create_policy_analysis_sheet(wb)
        self.create_scenario_modeling_sheet(wb)
        self.create_risk_assessment_sheet(wb)
        self.create_pivot_tables_sheet(wb)
        self.create_charts_sheet(wb)
        
        # Save workbook
        output_file = self.output_dir / f"trade_analysis_template_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(output_file)
        print(f"Excel template saved: {output_file}")
        
        return output_file
    
    def create_dashboard_sheet(self, wb):
        """Create main dashboard sheet."""
        ws = wb.create_sheet("Dashboard")
        
        # Title
        ws['A1'] = "Global Trade Analysis Dashboard"
        ws['A1'].font = Font(size=20, bold=True, color="1f77b4")
        ws.merge_cells('A1:H1')
        
        # Date
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10, italic=True)
        
        # Key metrics
        ws['A4'] = "Key Metrics"
        ws['A4'].font = Font(size=14, bold=True)
        
        # Load data for metrics
        conn = sqlite3.connect(self.db_path)
        
        # Total countries
        countries_count = pd.read_sql_query("SELECT COUNT(*) as count FROM countries", conn).iloc[0]['count']
        ws['A6'] = "Total Countries"
        ws['B6'] = countries_count
        
        # Total trade records
        trade_count = pd.read_sql_query("SELECT COUNT(*) as count FROM trade_data", conn).iloc[0]['count']
        ws['A7'] = "Trade Records"
        ws['B7'] = trade_count
        
        # Total economic indicators
        econ_count = pd.read_sql_query("SELECT COUNT(*) as count FROM economic_indicators", conn).iloc[0]['count']
        ws['A8'] = "Economic Indicators"
        ws['B8'] = econ_count
        
        # Active sanctions
        sanctions_count = pd.read_sql_query("SELECT COUNT(*) as count FROM sanctions WHERE status = 'active'", conn).iloc[0]['count']
        ws['A9'] = "Active Sanctions"
        ws['B9'] = sanctions_count
        
        conn.close()
        
        # Style metrics
        for cell in ['A6', 'A7', 'A8', 'A9']:
            ws[cell].font = Font(bold=True)
        
        # Summary table
        ws['A11'] = "Recent Trade Summary"
        ws['A11'].font = Font(size=14, bold=True)
        
        # Get recent trade data
        conn = sqlite3.connect(self.db_path)
        recent_trade = pd.read_sql_query("""
            SELECT c.country_name, td.year, td.trade_flow, SUM(td.value_usd) as total_value
            FROM trade_data td
            LEFT JOIN countries c ON td.reporter_country_id = c.country_id
            WHERE td.year >= 2022
            GROUP BY c.country_name, td.year, td.trade_flow
            ORDER BY td.year DESC, total_value DESC
            LIMIT 20
        """, conn)
        conn.close()
        
        # Add headers
        headers = ['Country', 'Year', 'Trade Flow', 'Total Value (USD)']
        for i, header in enumerate(headers):
            cell = ws.cell(row=12, column=i+1)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for i, row in enumerate(recent_trade.itertuples()):
            ws.cell(row=13+i, column=1).value = row.country_name
            ws.cell(row=13+i, column=2).value = row.year
            ws.cell(row=13+i, column=3).value = row.trade_flow
            ws.cell(row=13+i, column=4).value = row.total_value
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_trade_data_sheet(self, wb):
        """Create trade data sheet with pivot tables."""
        ws = wb.create_sheet("Trade Data")
        
        # Title
        ws['A1'] = "Trade Data Analysis"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Load trade data
        conn = sqlite3.connect(self.db_path)
        trade_data = pd.read_sql_query("""
            SELECT td.*, 
                   c1.country_name as reporter_country,
                   c2.country_name as partner_country
            FROM trade_data td
            LEFT JOIN countries c1 ON td.reporter_country_id = c1.country_id
            LEFT JOIN countries c2 ON td.partner_country_id = c2.country_id
            ORDER BY td.year DESC, td.value_usd DESC
        """, conn)
        conn.close()
        
        if not trade_data.empty:
            # Add data to sheet
            for r in dataframe_to_rows(trade_data, index=False, header=True):
                ws.append(r)
            
            # Style headers
            for cell in ws[3]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            
            # Create pivot table
            pivot_data = trade_data.pivot_table(
                index=['reporter_country', 'year'],
                columns='trade_flow',
                values='value_usd',
                aggfunc='sum'
            ).reset_index()
            
            # Add pivot table to sheet
            ws['A' + str(len(trade_data) + 5)] = "Trade Summary Pivot"
            ws['A' + str(len(trade_data) + 5)].font = Font(size=14, bold=True)
            
            pivot_headers = list(pivot_data.columns)
            for i, header in enumerate(pivot_headers):
                cell = ws.cell(row=len(trade_data) + 6, column=i+1)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for r in dataframe_to_rows(pivot_data, index=False, header=False):
                ws.append(r)
    
    def create_economic_indicators_sheet(self, wb):
        """Create economic indicators sheet."""
        ws = wb.create_sheet("Economic Indicators")
        
        # Title
        ws['A1'] = "Economic Indicators Analysis"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Load economic data
        conn = sqlite3.connect(self.db_path)
        economic_data = pd.read_sql_query("""
            SELECT ei.*, c.country_name
            FROM economic_indicators ei
            LEFT JOIN countries c ON ei.country_id = c.country_id
            ORDER BY ei.year DESC, ei.indicator_name
        """, conn)
        conn.close()
        
        if not economic_data.empty:
            # Add data to sheet
            for r in dataframe_to_rows(economic_data, index=False, header=True):
                ws.append(r)
            
            # Style headers
            for cell in ws[3]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            
            # Create pivot table for indicators
            pivot_data = economic_data.pivot_table(
                index=['country_name', 'year'],
                columns='indicator_name',
                values='indicator_value',
                aggfunc='mean'
            ).reset_index()
            
            # Add pivot table
            ws['A' + str(len(economic_data) + 5)] = "Economic Indicators Summary"
            ws['A' + str(len(economic_data) + 5)].font = Font(size=14, bold=True)
            
            pivot_headers = list(pivot_data.columns)
            for i, header in enumerate(pivot_headers):
                cell = ws.cell(row=len(economic_data) + 6, column=i+1)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for r in dataframe_to_rows(pivot_data, index=False, header=False):
                ws.append(r)
    
    def create_policy_analysis_sheet(self, wb):
        """Create policy analysis sheet."""
        ws = wb.create_sheet("Policy Analysis")
        
        # Title
        ws['A1'] = "Trade Policy Analysis"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Load policy data
        conn = sqlite3.connect(self.db_path)
        
        # Tariffs
        tariffs = pd.read_sql_query("""
            SELECT t.*, 
                   c1.country_name as imposing_country,
                   c2.country_name as target_country
            FROM tariffs t
            LEFT JOIN countries c1 ON t.country_id = c1.country_id
            LEFT JOIN countries c2 ON t.partner_country_id = c2.country_id
        """, conn)
        
        # Sanctions
        sanctions = pd.read_sql_query("""
            SELECT s.*, 
                   c1.country_name as sanctioning_country,
                   c2.country_name as target_country
            FROM sanctions s
            LEFT JOIN countries c1 ON s.sanctioning_country_id = c1.country_id
            LEFT JOIN countries c2 ON s.target_country_id = c2.country_id
        """, conn)
        
        conn.close()
        
        # Add tariffs data
        ws['A3'] = "Tariff Data"
        ws['A3'].font = Font(size=14, bold=True)
        
        if not tariffs.empty:
            for r in dataframe_to_rows(tariffs, index=False, header=True):
                ws.append(r)
            
            # Style headers
            for cell in ws[5]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Add sanctions data
        sanctions_start_row = len(tariffs) + 8
        ws['A' + str(sanctions_start_row)] = "Sanctions Data"
        ws['A' + str(sanctions_start_row)].font = Font(size=14, bold=True)
        
        if not sanctions.empty:
            for r in dataframe_to_rows(sanctions, index=False, header=True):
                ws.append(r)
            
            # Style headers
            for cell in ws[sanctions_start_row + 2]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    def create_scenario_modeling_sheet(self, wb):
        """Create scenario modeling sheet with financial models."""
        ws = wb.create_sheet("Scenario Modeling")
        
        # Title
        ws['A1'] = "Scenario Modeling & Financial Analysis"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Base scenario
        ws['A3'] = "Base Scenario (Current State)"
        ws['A3'].font = Font(size=14, bold=True)
        
        # Create base scenario table
        base_scenario_data = {
            'Metric': ['Total Trade Volume', 'Average Tariff Rate', 'GDP Growth', 'Unemployment Rate'],
            'Value': [1000000, 2.5, 3.2, 5.1],
            'Unit': ['USD', '%', '%', '%']
        }
        
        base_df = pd.DataFrame(base_scenario_data)
        
        # Add headers
        headers = ['Metric', 'Value', 'Unit']
        for i, header in enumerate(headers):
            cell = ws.cell(row=4, column=i+1)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for r in dataframe_to_rows(base_df, index=False, header=False):
            ws.append(r)
        
        # Scenario analysis
        scenario_start_row = 10
        ws['A' + str(scenario_start_row)] = "Scenario Analysis"
        ws['A' + str(scenario_start_row)].font = Font(size=14, bold=True)
        
        # Create scenario comparison table
        scenarios = ['Base', 'Optimistic', 'Pessimistic', 'Tariff Increase', 'Trade Agreement']
        metrics = ['Trade Volume', 'GDP Impact', 'Employment Impact', 'Risk Score']
        
        # Headers
        for i, scenario in enumerate(scenarios):
            cell = ws.cell(row=scenario_start_row + 1, column=i+2)
            cell.value = scenario
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Metric labels
        for i, metric in enumerate(metrics):
            cell = ws.cell(row=scenario_start_row + 2 + i, column=1)
            cell.value = metric
            cell.font = Font(bold=True)
        
        # Scenario values (simplified)
        scenario_values = [
            [100, 0, 0, 50],      # Base
            [120, 2.5, -0.5, 30], # Optimistic
            [80, -2.0, 1.0, 70],  # Pessimistic
            [85, -1.5, 0.8, 65],  # Tariff Increase
            [115, 1.8, -0.3, 35]  # Trade Agreement
        ]
        
        for i, values in enumerate(scenario_values):
            for j, value in enumerate(values):
                cell = ws.cell(row=scenario_start_row + 2 + i, column=j+2)
                cell.value = value
        
        # Financial modeling section
        model_start_row = scenario_start_row + 8
        ws['A' + str(model_start_row)] = "Financial Modeling"
        ws['A' + str(model_start_row)].font = Font(size=14, bold=True)
        
        # Create cash flow model
        years = range(2024, 2030)
        cash_flow_data = {
            'Year': years,
            'Revenue': [1000000 * (1.05 ** i) for i in range(len(years))],
            'Costs': [800000 * (1.03 ** i) for i in range(len(years))],
            'Net Income': [],
            'Cumulative Cash Flow': []
        }
        
        cumulative = 0
        for i in range(len(years)):
            net_income = cash_flow_data['Revenue'][i] - cash_flow_data['Costs'][i]
            cash_flow_data['Net Income'].append(net_income)
            cumulative += net_income
            cash_flow_data['Cumulative Cash Flow'].append(cumulative)
        
        cash_flow_df = pd.DataFrame(cash_flow_data)
        
        # Add cash flow table
        for r in dataframe_to_rows(cash_flow_df, index=False, header=True):
            ws.append(r)
        
        # Style headers
        for cell in ws[model_start_row + 2]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    def create_risk_assessment_sheet(self, wb):
        """Create risk assessment sheet."""
        ws = wb.create_sheet("Risk Assessment")
        
        # Title
        ws['A1'] = "Risk Assessment & Scoring"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Risk factors
        ws['A3'] = "Risk Factors"
        ws['A3'].font = Font(size=14, bold=True)
        
        risk_factors = [
            ['Trade Volatility', 25, 'High', 'Medium', 'Low'],
            ['Political Stability', 20, 'Low', 'Medium', 'High'],
            ['Economic Growth', 20, 'Low', 'Medium', 'High'],
            ['Regulatory Environment', 15, 'Unfavorable', 'Neutral', 'Favorable'],
            ['Geographic Risk', 10, 'High', 'Medium', 'Low'],
            ['Currency Risk', 10, 'High', 'Medium', 'Low']
        ]
        
        # Headers
        headers = ['Risk Factor', 'Weight (%)', 'High Risk', 'Medium Risk', 'Low Risk']
        for i, header in enumerate(headers):
            cell = ws.cell(row=4, column=i+1)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add risk factors
        for i, factor in enumerate(risk_factors):
            for j, value in enumerate(factor):
                cell = ws.cell(row=5+i, column=j+1)
                cell.value = value
        
        # Risk scoring matrix
        scoring_start_row = 15
        ws['A' + str(scoring_start_row)] = "Risk Scoring Matrix"
        ws['A' + str(scoring_start_row)].font = Font(size=14, bold=True)
        
        # Create scoring matrix
        countries = ['USA', 'China', 'Germany', 'Japan', 'UK']
        risk_scores = [35, 65, 25, 30, 40]
        
        # Headers
        ws.cell(row=scoring_start_row + 1, column=1).value = "Country"
        ws.cell(row=scoring_start_row + 1, column=2).value = "Risk Score"
        ws.cell(row=scoring_start_row + 1, column=3).value = "Risk Level"
        
        for cell in ws[scoring_start_row + 1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Add country data
        for i, (country, score) in enumerate(zip(countries, risk_scores)):
            ws.cell(row=scoring_start_row + 2 + i, column=1).value = country
            ws.cell(row=scoring_start_row + 2 + i, column=2).value = score
            
            # Risk level
            if score < 30:
                risk_level = "Low"
            elif score < 60:
                risk_level = "Medium"
            else:
                risk_level = "High"
            
            ws.cell(row=scoring_start_row + 2 + i, column=3).value = risk_level
    
    def create_pivot_tables_sheet(self, wb):
        """Create pivot tables sheet."""
        ws = wb.create_sheet("Pivot Tables")
        
        # Title
        ws['A1'] = "Pivot Tables & Data Analysis"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Load data for pivot tables
        conn = sqlite3.connect(self.db_path)
        
        # Trade data for pivots
        trade_pivot_data = pd.read_sql_query("""
            SELECT td.year, td.trade_flow, td.value_usd,
                   c1.country_name as reporter_country,
                   c2.country_name as partner_country
            FROM trade_data td
            LEFT JOIN countries c1 ON td.reporter_country_id = c1.country_id
            LEFT JOIN countries c2 ON td.partner_country_id = c2.country_id
            WHERE td.year >= 2020
        """, conn)
        
        conn.close()
        
        if not trade_pivot_data.empty:
            # Pivot 1: Trade by country and year
            pivot1 = trade_pivot_data.pivot_table(
                index='reporter_country',
                columns='year',
                values='value_usd',
                aggfunc='sum'
            ).reset_index()
            
            ws['A3'] = "Trade Volume by Country and Year"
            ws['A3'].font = Font(size=14, bold=True)
            
            for r in dataframe_to_rows(pivot1, index=False, header=True):
                ws.append(r)
            
            # Style headers
            for cell in ws[5]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Pivot 2: Trade flow analysis
            pivot2 = trade_pivot_data.pivot_table(
                index='reporter_country',
                columns='trade_flow',
                values='value_usd',
                aggfunc='sum'
            ).reset_index()
            
            pivot2_start_row = len(pivot1) + 8
            ws['A' + str(pivot2_start_row)] = "Trade Flow Analysis (Imports vs Exports)"
            ws['A' + str(pivot2_start_row)].font = Font(size=14, bold=True)
            
            for r in dataframe_to_rows(pivot2, index=False, header=True):
                ws.append(r)
            
            # Style headers
            for cell in ws[pivot2_start_row + 2]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    def create_charts_sheet(self, wb):
        """Create charts sheet with visualizations."""
        ws = wb.create_sheet("Charts")
        
        # Title
        ws['A1'] = "Data Visualizations"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Load data for charts
        conn = sqlite3.connect(self.db_path)
        
        # Trade trends data
        trade_trends = pd.read_sql_query("""
            SELECT td.year, td.trade_flow, SUM(td.value_usd) as total_value
            FROM trade_data td
            WHERE td.year >= 2020
            GROUP BY td.year, td.trade_flow
            ORDER BY td.year, td.trade_flow
        """, conn)
        
        conn.close()
        
        if not trade_trends.empty:
            # Add data for chart
            ws['A3'] = "Trade Trends Data"
            ws['A3'].font = Font(size=14, bold=True)
            
            for r in dataframe_to_rows(trade_trends, index=False, header=True):
                ws.append(r)
            
            # Style headers
            for cell in ws[5]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Create chart
            chart = LineChart()
            chart.title = "Global Trade Trends"
            chart.x_axis.title = "Year"
            chart.y_axis.title = "Trade Value (USD)"
            
            data = Reference(ws, min_col=3, min_row=4, max_row=len(trade_trends)+4, max_col=3)
            cats = Reference(ws, min_col=1, min_row=5, max_row=len(trade_trends)+4)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, "E3")

def main():
    """Main function to generate Excel template."""
    try:
        generator = ExcelTemplateGenerator()
        output_file = generator.generate_template()
        print(f"Excel template generated successfully: {output_file}")
        
    except Exception as e:
        print(f"Error generating Excel template: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main() 